"""
Сервисные функции для импорта Excel и расчетов отчетов.
"""

from __future__ import annotations

import difflib
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook

from django.db import transaction
from django.db.models import Avg, Count

from .models import ArrivalEvent, Bull, ExcelImportBatch, ExcelImportPendingRow, WeightRecord


def split_external_id(external_id: str) -> tuple[str, str]:
    """
    Делит полный номер из файла на две части:
    - первые 8 символов (бивайка),
    - последние 6 символов (номер быка).
    """

    normalized = (external_id or "").strip()
    if len(normalized) < 14:
        return normalized[:8], normalized[-6:]
    return normalized[:8], normalized[-6:]


def get_similar_ids(raw_external_id: str, limit: int = 5) -> list[str]:
    """
    Возвращает список похожих номеров быков для исправления опечаток.

    Используется стандартный `difflib`, которого достаточно для учебного
    проекта и коротких строковых идентификаторов.
    """

    all_ids = list(Bull.objects.values_list("external_id", flat=True))
    if not all_ids:
        return []
    return difflib.get_close_matches(raw_external_id, all_ids, n=limit, cutoff=0.55)


@dataclass
class ImportResult:
    """
    Результат выполнения импорта Excel.
    """

    batch: ExcelImportBatch
    applied_rows: int
    pending_rows: int


@transaction.atomic
def import_weights_excel(file_obj) -> ImportResult:
    """
    Импортирует рабочий Excel-файл с весами.

    Формат строго соответствует промту:
    - колонка A: полный номер быка;
    - заголовки колонок B, C, D ...: даты взвешиваний;
    - значения под датами: вес быка на соответствующую дату.

    Если номер не найден в базе, строка не применяется автоматически.
    Она попадает в список проблемных строк с похожими номерами, потому что
    по требованию пользователя это может быть опечатка в номере.
    """

    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    weighing_dates = _read_weighing_dates_from_header(ws)
    if not weighing_dates:
        raise ValueError("В Excel не найдены даты взвешиваний в заголовках колонок B, C, D ...")

    batch = ExcelImportBatch.objects.create(
        previous_weighing_date=weighing_dates[0],
        current_weighing_date=weighing_dates[-1],
        weighing_dates=[d.isoformat() for d in weighing_dates],
    )

    applied = 0
    pending = 0
    total = 0

    for row_idx in range(2, ws.max_row + 1):
        raw_external_id = str(ws.cell(row=row_idx, column=1).value or "").strip()

        if not raw_external_id:
            continue

        weights_by_date = {}
        for col_offset, weighing_date in enumerate(weighing_dates, start=2):
            weight_value = ws.cell(row=row_idx, column=col_offset).value
            if weight_value is None or weight_value == "":
                continue
            weights_by_date[weighing_date.isoformat()] = str(_as_decimal(weight_value))

        if not weights_by_date:
            continue

        total += 1
        bull = Bull.objects.filter(external_id=raw_external_id).first()

        if bull:
            _apply_weight_map(bull, weights_by_date)
            applied += 1
            continue

        suggested = get_similar_ids(raw_external_id)
        ordered_weights = list(weights_by_date.values())
        ExcelImportPendingRow.objects.create(
            batch=batch,
            raw_external_id=raw_external_id,
            previous_weight_kg=_as_decimal(ordered_weights[0]),
            current_weight_kg=_as_decimal(ordered_weights[-1]),
            weights_by_date=weights_by_date,
            suggested_ids=suggested,
        )
        pending += 1

    batch.total_rows = total
    batch.applied_rows = applied
    batch.pending_rows = pending
    batch.save(update_fields=["total_rows", "applied_rows", "pending_rows"])

    return ImportResult(batch=batch, applied_rows=applied, pending_rows=pending)


def _read_weighing_dates_from_header(ws) -> list[date]:
    """
    Читает даты взвешиваний из первой строки Excel.

    Пользователь описал формат так: A - номер быка, B/C/D... - даты.
    Поэтому мы идем по заголовкам начиная со второй колонки и пытаемся
    превратить каждое значение в дату Python.
    """

    dates = []
    for col_idx in range(2, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col_idx).value
        if header_value in (None, ""):
            continue
        dates.append(_parse_excel_date_header(header_value))
    return dates


def _parse_excel_date_header(value) -> date:
    """
    Преобразует заголовок Excel в дату.

    Поддерживаются настоящие Excel-даты и распространенные текстовые форматы,
    например `25.04.2026` или `2026-04-25`.
    """

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    raw_value = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw_value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Не удалось распознать дату в заголовке Excel: {raw_value}")


def _as_decimal(value) -> Decimal:
    """
    Приводит значение из Excel к Decimal.
    """

    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value).replace(",", "."))


def _apply_weight_pair(bull, previous_date, current_date, prev_weight, curr_weight) -> None:
    """
    Создает или обновляет пару записей веса (прошлая и текущая дата).
    """

    WeightRecord.objects.update_or_create(
        bull=bull,
        weighing_date=previous_date,
        source="excel",
        defaults={"weight_kg": _as_decimal(prev_weight), "note": "Импорт: предыдущий вес"},
    )


def _apply_weight_map(bull: Bull, weights_by_date: dict[str, str]) -> None:
    """
    Создает или обновляет все записи веса из одной строки Excel.

    Словарь приходит в виде `{YYYY-MM-DD: вес}`. Такая структура нужна,
    потому что в реальном файле может быть не две, а несколько дат перевески.
    """

    for date_text, weight_value in weights_by_date.items():
        WeightRecord.objects.update_or_create(
            bull=bull,
            weighing_date=datetime.strptime(date_text, "%Y-%m-%d").date(),
            source="excel",
            defaults={"weight_kg": _as_decimal(weight_value), "note": "Импорт из Excel"},
        )
    WeightRecord.objects.update_or_create(
        bull=bull,
        weighing_date=current_date,
        source="excel",
        defaults={"weight_kg": _as_decimal(curr_weight), "note": "Импорт: текущий вес"},
    )


@transaction.atomic
def resolve_pending_row(pending_row: ExcelImportPendingRow, bull: Bull) -> None:
    """
    Применяет проблемную строку импорта к выбранному быку.
    """

    if pending_row.is_resolved:
        return

    batch = pending_row.batch
    if pending_row.weights_by_date:
        _apply_weight_map(bull, pending_row.weights_by_date)
    else:
        _apply_weight_pair(
            bull=bull,
            previous_date=batch.previous_weighing_date,
            current_date=batch.current_weighing_date,
            prev_weight=pending_row.previous_weight_kg,
            curr_weight=pending_row.current_weight_kg,
        )

    pending_row.is_resolved = True
    pending_row.resolved_to = bull
    pending_row.save(update_fields=["is_resolved", "resolved_to"])

    batch.applied_rows += 1
    batch.pending_rows = max(batch.pending_rows - 1, 0)
    batch.save(update_fields=["applied_rows", "pending_rows"])


@transaction.atomic
def create_bull_from_pending_row(pending_row: ExcelImportPendingRow, arrived_at, section, comment: str = "") -> Bull:
    """
    Создает нового быка из проблемной строки импорта.

    Это ручной шаг: если пользователь видит, что неизвестный номер не опечатка,
    а новый бык, он выбирает секцию поступления, и система создает карточку,
    событие поступления и применяет веса из Excel.
    """

    if pending_row.is_resolved:
        return pending_row.resolved_to

    bull = Bull.objects.create(external_id=pending_row.raw_external_id, section=section)
    ArrivalEvent.objects.create(bull=bull, arrived_at=arrived_at, section=section, comment=comment)
    resolve_pending_row(pending_row, bull)
    return bull


def build_dashboard_stats() -> dict:
    """
    Собирает агрегаты для главной панели и простых отчетов.
    """

    section_counts = (
        Bull.objects.filter(is_active=True, section__isnull=False)
        .values("section__name")
        .annotate(total=Count("id"), avg_weight=Avg("weight_records__weight_kg"))
        .order_by("section__name")
    )

    return {
        "total_active_bulls": Bull.objects.filter(is_active=True).count(),
        "section_counts": list(section_counts),
    }


def build_growth_rating(limit: int = 10) -> dict:
    """
    Формирует рейтинги слабого и лучшего набора веса.

    Метрики:
    - абсолютный привес = последний вес - первый вес;
    - среднесуточный привес = абсолютный / количество дней между замерами.
    """

    bulls_data = []
    for bull in Bull.objects.filter(is_active=True).prefetch_related("weight_records"):
        records = list(bull.weight_records.order_by("weighing_date"))
        if len(records) < 2:
            continue

        first = records[0]
        last = records[-1]
        days = (last.weighing_date - first.weighing_date).days or 1
        absolute_gain = Decimal(last.weight_kg) - Decimal(first.weight_kg)
        daily_gain = absolute_gain / Decimal(days)

        bulls_data.append(
            {
                "bull": bull,
                "absolute_gain": absolute_gain,
                "daily_gain": daily_gain,
                "days": days,
                "first_weight": first.weight_kg,
                "last_weight": last.weight_kg,
            }
        )

    by_absolute = sorted(bulls_data, key=lambda x: x["absolute_gain"])
    by_daily = sorted(bulls_data, key=lambda x: x["daily_gain"])

    return {
        "weak_absolute": by_absolute[:limit],
        "best_absolute": list(reversed(by_absolute[-limit:])),
        "weak_daily": by_daily[:limit],
        "best_daily": list(reversed(by_daily[-limit:])),
    }
